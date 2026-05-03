from pathlib import Path
from datetime import datetime, time, timedelta
import csv
import re
import unicodedata
from openpyxl import load_workbook


# Files
MAPPING_CSV = Path("calc_simulator_and_corresponding_physiological_files.csv")
CAT_CSV = Path("combined_cat_class.csv")

# Output
OUTPUT_SUFFIX = "_completed_with_categories.xlsx"

# Path correction: H drive -> G drive
OLD_PATH = r"H:\My Drive\Ariel Uni"
NEW_PATH = r"G:\My Drive\Ariel Uni"

"""Normalize text for"""
def clean_text(value):
    if value is None:
        return ""

    value = str(value)
    value = unicodedata.normalize("NFKC", value)
    value = value.replace("\u200f", "").replace("\u200e", "").replace("\ufeff", "")
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)

    return value.strip()

"""Replace H drive with G drive."""
def fix_path(path):
    path = clean_text(path).strip('"')

    if path.lower().startswith(OLD_PATH.lower()):
        path = NEW_PATH + path[len(OLD_PATH):]

    return path

"""Convert time to HH:MM:SS format."""
def clean_time(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.time()

    if isinstance(value, time):
        return f"{value.hour:02}:{value.minute:02}:{value.second:02}"

    if isinstance(value, timedelta):
        seconds = round(value.total_seconds())
        return seconds_to_time(seconds)

    if isinstance(value, (int, float)):
        if 0 <= value < 1:
            seconds = round(value * 24 * 3600)  # Excel time format
        else:
            seconds = round(value)

        return seconds_to_time(seconds)

    value = str(value).strip().replace(".", ":")
    parts = value.split(":")

    try:
        if len(parts) == 3:
            h, m, s = parts
            return f"{int(h):02}:{int(m):02}:{int(float(s)):02}"

        if len(parts) == 2:
            h, m = parts
            return f"{int(h):02}:{int(m):02}:00"

    except ValueError:
        pass

    return value
"""Convert seconds to HH:MM:SS."""
def seconds_to_time(seconds):

    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60

    return f"{h:02}:{m:02}:{s:02}"

"""read CSV File"""
def read_csv_file(path):
    with open(path, "r", encoding="cp1255", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return reader.fieldnames, rows

"""Find time and text colums"""
def find_time_and_text_columns(ws):

    for row in range(1, min(ws.max_row, 30) + 1):
        headers = {}

        for col in range(1, ws.max_column + 1):
            header = clean_text(ws.cell(row=row, column=col).value)

            if header:
                headers[header] = col

        if "זמן" in headers and "טקסט" in headers:
            return row, headers["זמן"], headers["טקסט"]

    raise ValueError("Could not find 'זמן' and 'טקסט' columns")

"""Checking strick of columns is empty"""
def column_is_empty(ws, col):
    """Check if an entire Excel column is empty."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return False

    return True

"""Check if colum is empty"""
def find_empty_columns(ws, number_of_columns):
    """Find the first block of empty columns."""
    for start_col in range(1, ws.max_column + number_of_columns + 50):
        empty_block = True

        for col in range(start_col, start_col + number_of_columns):
            if not column_is_empty(ws, col):
                empty_block = False
                break

        if empty_block:
            return start_col

    raise ValueError("Could not find empty columns")


"""Build table from combined_cat_class.csv.""" # (זמן, טקסט)  ->  ערכי הקטגוריות
def make_category_lookup(cat_rows, cat_headers, filename):

    category_headers = cat_headers[3:14]  # CSV columns D:N
    lookup = {}

    for row in cat_rows:
        if clean_text(row.get("filename")) != clean_text(filename):
            continue

        key = (
            clean_time(row.get("time_shira")),
            clean_text(row.get("text_shira")),
        )

        values = []
        for header in category_headers:
            value = row.get(header)

            if value is None or str(value).strip() == "":
                values.append(None)
            elif str(value).strip().isdigit():
                values.append(int(value))
            else:
                values.append(value)

        lookup[key] = values

    return category_headers, lookup

"""Add category columns to each participant transcript file."""
def process_transcript(part_xlsx_path, cat_rows, cat_headers):
    # Convert the input path string to a Path object
    part_xlsx_path = Path(part_xlsx_path)

    # Check if the Excel file exists
    if not part_xlsx_path.exists():
        print(f"[SKIP] File not found: {part_xlsx_path}")
        return False

    # Create a new output file path in the same folder as the original file
    output_path = part_xlsx_path.with_name(part_xlsx_path.stem + OUTPUT_SUFFIX)

    # Safety check: make sure the output file is not the same as the original file
    if output_path.resolve() == part_xlsx_path.resolve():
        raise RuntimeError("Output path is identical to source path")

    # Safety check: do not overwrite an existing output file
    if output_path.exists():
        raise FileExistsError(f"Output already exists, not overwriting: {output_path}")

    # Open the participent file workbook
    wb = load_workbook(part_xlsx_path)

    # Select the active worksheet
    ws = wb.active
    
    # Find the header row and the columns of "זמן" and "טקסט"
    header_row, time_col, text_col = find_time_and_text_columns(ws)

    # Build a lookup table from the category cat CSV for this specific Excel file
    category_headers, lookup = make_category_lookup(
        cat_rows,        # All rows from combined_cat_class.csv
        cat_headers,     # Column names from combined_cat_class.csv
        part_xlsx_path.name,  # Current Excel filename
    )

    # Find the first empty block of columns where the categories can be paste
    start_col = find_empty_columns(ws, len(category_headers))

    # Write the category column names into the Excel header row
    for i, header in enumerate(category_headers):
        ws.cell(row=header_row, column=start_col + i).value = header

    # Counter for the number of matched rows
    matched = 0

    # Go over all data rows in the Excel file, starting after the header row
    for row in range(header_row + 1, ws.max_row + 1):

        # Clear the cell data "time" "text" before using
        key = (
            clean_time(ws.cell(row=row, column=time_col).value),
            clean_text(ws.cell(row=row, column=text_col).value),
        )

        # If this time + text does not exist in the category lookup, skip this row
        if key not in lookup:
            continue

        # If a match exists, write the category values into the new columns
        for i, value in enumerate(lookup[key]):
            ws.cell(row=row, column=start_col + i).value = value

        # Count this row as matched
        matched += 1

    # Save the modified workbook as a new file, not over the original file
    wb.save(output_path)

    print(f"Saved: {output_path}")

    print(f"Matched rows: {matched}")

    return True

"""Add category columns to each path participent file file."""
def main():
    # Read input files
    _, mapping_rows = read_csv_file(MAPPING_CSV)
    cat_headers, cat_rows = read_csv_file(CAT_CSV)

    # Collect all transcript paths from mapping table
    transcript_paths = []

    for row in mapping_rows:
        path = fix_path(row.get("TranscriptionManualFile"))

        if path and path not in transcript_paths:
            transcript_paths.append(path)

    print(f"Transcript files found: {len(transcript_paths)}")

    # Process each participant file
    processed = 0

    for path in transcript_paths:
        try:
            if process_transcript(path, cat_rows, cat_headers):
                processed += 1

        except Exception as e:
            print(f"[ERROR] {path}")
            print(e)

    print(f"Done. Processed files: {processed}")


if __name__ == "__main__":
    main()
